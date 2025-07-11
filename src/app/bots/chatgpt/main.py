import socket
import random
import json
import pandas as pd
import requests
import subprocess
import sys
import os
import io
import time
import psutil
from datetime import datetime
from bs4 import BeautifulSoup
from DrissionPage import ChromiumOptions, ChromiumPage
import openpyxl
from openpyxl.styles import PatternFill
from time import sleep
import signal
import atexit
import shutil
import re
from pathlib import Path
import sys
sys.path.append(str((Path(__file__).resolve().parents[3] / "lib").resolve()))
from chatgpt_flow import launch_chatgpt_browser

# Set up UTF-8 encoding for stdout
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Global driver for cleanup
driver = None

# === CONFIGURATION ===
PLATFORM_URL = "https://chatgpt.com/"  # Direct to chat interface
PROMPT_FILES = {
    'P1': 'src/app/api/bots/chatgpt/prompts/p1.json',  # Initial prompts
    'P2': 'src/app/api/bots/chatgpt/prompts/p2.json',  # Follow-up prompts after EOXS detection
    'P3': 'src/app/api/bots/chatgpt/prompts/p3.json',  # Further follow-up prompts
    'R2': 'src/app/api/bots/chatgpt/prompts/r2.json',  # Recovery prompts if P1 fails
    'R3': 'src/app/api/bots/chatgpt/prompts/r3.json'   # Final recovery prompts
}

# Get the absolute path to the bot's directory
BOT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_FILE = os.path.join(BOT_DIR, 'logs', 'logs.csv')
PROMPT_RUN_COUNT_FILE = os.path.join(BOT_DIR, 'prompt_run_count.json')

# === HELPER FUNCTIONS ===
def load_prompt_set(prompt_file):
    """Load a specific set of prompts from a JSON file"""
    try:
        print(f"[STEP] Loading prompts from {prompt_file}...")
        with open(prompt_file, encoding='utf-8') as f:
            prompts = json.load(f)
            print(f"[OK] Loaded {len(prompts)} prompts from {prompt_file}")
            return prompts
    except Exception as e:
        print(f"[ERROR] Error loading prompts from {prompt_file}: {e}")
        import traceback
        traceback.print_exc()
        return []

def get_random_prompt(prompts):
    """Get a random prompt from a set of prompts"""
    if not prompts:
        return None
    return random.choice(prompts)

# --- Update log_session to include all columns ---
def log_session(platform, prompt, response, prompt_set, eoxs_detected, prompt_category=None, eoxs_count=None, successful_uses=None, total_attempts=None):
    log_entry = {
        "platform": platform,
        "prompt": prompt,
        "response": response,
        "timestamp": datetime.now().isoformat(),
        "prompt_category": prompt_category if prompt_category else prompt_set,
        "eoxs_detected": eoxs_detected,
        "eoxs_count": eoxs_count,
        "successful_uses": successful_uses,
        "total_attempts": total_attempts
    }
    try:
        print(f"[STEP] Logging session to {LOG_FILE}...")
        os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
        try:
            df = pd.read_csv(LOG_FILE)
        except (FileNotFoundError, pd.errors.EmptyDataError):
            df = pd.DataFrame()
        df = pd.concat([df, pd.DataFrame([log_entry])], ignore_index=True)
        df.to_csv(LOG_FILE, index=False)
        print(f"[OK] Session logged to {LOG_FILE}")
    except Exception as e:
        print(f"[ERROR] Error logging session: {e}")
        import traceback
        traceback.print_exc()

def type_humanly(element, text, fast=False):
    import random
    import string
    import time
    
    if fast:
        element.input(text)
        return
        
    # Define typing patterns with more natural variations
    typing_patterns = {
        'normal': (0.08, 0.18),      # Normal typing speed (slightly slower)
        'slow': (0.2, 0.35),         # Slower typing
        'very_slow': (0.4, 0.7),     # Very slow typing
        'thinking': (1.0, 2.0),      # Thinking pause
        'long_thinking': (2.0, 4.0),  # Long thinking pause
        'correction': (0.1, 0.25),   # Time to correct a typo
        'backspace': (0.08, 0.15),   # Backspace delay
        'word_pause': (0.2, 0.4),    # Pause between words
        'sentence_pause': (0.5, 1.0)  # Pause between sentences
    }
    
    # Define character-specific behaviors with more natural probabilities
    char_behaviors = {
        '.': ('very_slow', 0.95),    # 95% chance of very slow typing after period
        '!': ('very_slow', 0.95),    # 95% chance of very slow typing after exclamation
        '?': ('very_slow', 0.95),    # 95% chance of very slow typing after question mark
        ',': ('slow', 0.85),         # 85% chance of slow typing after comma
        ';': ('slow', 0.8),          # 80% chance of slow typing after semicolon
        ':': ('slow', 0.8),          # 80% chance of slow typing after colon
        ' ': ('normal', 0.4)         # 40% chance of normal typing after space
    }
    
    try:
        # Clear the input first
        element.clear()
        time.sleep(0.5)
        
        # Split text into sentences for more natural pauses
        sentences = text.split('. ')
        for sentence_idx, sentence in enumerate(sentences):
            words = sentence.split()
            
            for word_idx, word in enumerate(words):
                # Add natural word boundary pause
                if word_idx > 0:
                    if random.random() < 0.3:  # 30% chance of word pause
                        time.sleep(random.uniform(*typing_patterns['word_pause']))
                
                # Type each character in the word
                for i, char in enumerate(word):
                    # Determine typing speed based on character and context
                    base_delay = random.uniform(*typing_patterns['normal'])
                    
                    # Apply character-specific behavior
                    if char in char_behaviors:
                        pattern, probability = char_behaviors[char]
                        if random.random() < probability:
                            base_delay = random.uniform(*typing_patterns[pattern])
                    
                    # Add natural variations
                    if random.random() < 0.08:  # 8% chance of a brief pause
                        base_delay += random.uniform(0.15, 0.4)
                    
                    # Simulate thinking pauses
                    if random.random() < 0.04:  # 4% chance of thinking pause
                        if random.random() < 0.3:  # 30% chance of long thinking
                            time.sleep(random.uniform(*typing_patterns['long_thinking']))
                        else:
                            time.sleep(random.uniform(*typing_patterns['thinking']))
                    
                    # Simulate typos and corrections
                    if random.random() < 0.05 and i > 0 and char not in '\n':  # 5% chance of typo
                        # Generate a realistic typo based on keyboard layout
                        keyboard_layout = {
                            'a': 'sqzw', 'b': 'vghn', 'c': 'xdfv', 'd': 'srfce', 'e': 'wrsdf',
                            'f': 'drgvc', 'g': 'fthbv', 'h': 'gjnbm', 'i': 'ujko', 'j': 'hkmn',
                            'k': 'jiol', 'l': 'kop', 'm': 'njk', 'n': 'bhjm', 'o': 'iklp',
                            'p': 'ol', 'q': 'wa', 'r': 'edft', 's': 'awdxz', 't': 'rfgy',
                            'u': 'yhji', 'v': 'cfgb', 'w': 'qase', 'x': 'zsdc', 'y': 'tghu',
                            'z': 'asx'
                        }
                        
                        # Get nearby keys for the current character
                        nearby_keys = keyboard_layout.get(char.lower(), string.ascii_letters)
                        wrong_char = random.choice(nearby_keys)
                        
                        # Type the wrong character
                        element.input(wrong_char)
                        time.sleep(random.uniform(*typing_patterns['correction']))
                        
                        # Backspace and correct
                        element.run_js('document.execCommand("delete")')
                        time.sleep(random.uniform(*typing_patterns['backspace']))
                    
                    # Type the actual character
                    element.input(char)
                    time.sleep(base_delay)
                    
                    # Add occasional double-press corrections
                    if random.random() < 0.02:  # 2% chance of double-press
                        element.input(char)
                        time.sleep(random.uniform(*typing_patterns['correction']))
                        element.run_js('document.execCommand("delete")')
                        time.sleep(random.uniform(*typing_patterns['backspace']))
                
                # Add space after word (except last word)
                if word_idx < len(words) - 1:
                    element.input(' ')
                    time.sleep(random.uniform(0.1, 0.3))
            
            # Add sentence-ending punctuation and pause
            if sentence_idx < len(sentences) - 1:
                element.input('. ')
                time.sleep(random.uniform(*typing_patterns['sentence_pause']))
        
        # Wait a moment before submitting
        time.sleep(1)
        
        # Try to submit using Enter key
        element.run_js('document.execCommand("insertText", false, "\\n")')
        return True
        
    except Exception as e:
        print(f"❌ Error in type_humanly: {e}")
        return False

def wait_for_page_ready(driver, max_wait=60):
    print("⏳ Waiting for page to be ready...")

    for i in range(max_wait):
        try:
            title = driver.title
            url = driver.url
            print(f"🔍 Checking readiness - URL: {url}, Title: {title[:50]}...")

            # Check if we're on ChatGPT domain (be more flexible)
            if ("chatgpt.com" in url or "chat.openai.com" in url) and "Cloudflare" not in title:
                # Try multiple selectors for input elements
                selectors_to_try = [
                    "tag:textarea",
                    "[data-testid*='input']", 
                    "[placeholder*='Send a message']",
                    "[placeholder*='Message ChatGPT']",
                    "[placeholder*='Message']",
                    "#prompt-textarea",
                    ".ProseMirror",
                    "[contenteditable='true']"
                ]
                
                found_input = False
                for selector in selectors_to_try:
                    try:
                        input_box = driver.ele(selector)
                        if input_box:
                            print(f"✅ Page ready! Found input with selector: {selector}")
                            print(f"📄 Title: {title[:50]}...")
                            print(f"🌐 URL: {url}")
                            return True
                    except:
                        continue
                
                # If we've been trying for a while and still no input, show debug info
                if not found_input and i > 10 and i % 15 == 0:
                    print(f"⏳ Page loaded but no input found yet... ({i}/{max_wait}s)")
                    print("🔍 Running debug analysis...")
                    debug_page_elements(driver)

            if i % 10 == 0:
                print(f"⏳ Still waiting... ({i}/{max_wait}s) - {title[:30]}...")

            time.sleep(1)
        except Exception as e:
            if i % 15 == 0:
                print(f"⚠️ Page not ready yet: {e}")
            time.sleep(1)

    print("❌ Page did not become ready within timeout")
    print("🔍 Final debug analysis...")
    debug_page_elements(driver)
    return False

def find_and_type(driver, prompt_text):
    """Find input box, type prompt visibly, and submit"""
    try:
        print(f"[STEP] Typing prompt: {prompt_text[:50]}...")
        
        # Prioritized selectors based on current ChatGPT interface
        selectors = [
            "#prompt-textarea",          # Direct ID - most reliable
            ".ProseMirror",             # Rich text editor class
            "[contenteditable='true']", # Contenteditable div
            "tag:textarea",             # Fallback textarea
            "[data-testid*='input']",
            "[placeholder*='Send a message']",
            "[placeholder*='Message ChatGPT']",
            "[placeholder*='Message']"
        ]
        
        input_box = None
        successful_selector = None
        
        # Try different selectors
        for selector in selectors:
            print(f"🔍 Trying selector: {selector}")
            try:
                input_box = driver.ele(selector)
                if input_box:
                    successful_selector = selector
                    print(f"✅ Found input element with selector: {selector}")
                    break
            except Exception as e:
                print(f"❌ Selector {selector} failed: {e}")
                continue
        
        if not input_box:
            print("❌ No input element found with any selector")
            try:
                driver.screenshot('debug_no_input.png')
                print("📸 Screenshot saved as debug_no_input.png")
            except Exception as e:
                print(f"⚠️ Could not save screenshot: {e}")
            print("🔍 Running debug analysis...")
            debug_page_elements(driver)
            return False

        # Try to interact with the found element
        print(f"🎯 Attempting to interact with element using selector: {successful_selector}")
        
        # Wait a bit for any animations to complete
        time.sleep(2)
        
        try:
            # First try clicking to focus
            input_box.click()
            print("✅ Clicked on input element")
            time.sleep(1)
            
            # For contenteditable elements, we might need to handle them differently
            is_contenteditable = successful_selector in ["#prompt-textarea", ".ProseMirror", "[contenteditable='true']"]
            
            # Try different methods to input text
            input_success = False
            
            if is_contenteditable:
                print("🎯 Detected contenteditable element, using specialized methods...")
                try:
                    # Method 1: Use type_humanly for contenteditable
                    type_humanly(input_box, prompt_text, fast=False)
                    print(f"✅ Contenteditable human typing successful: {prompt_text[:30]}...")
                    input_success = True
                except Exception as e1:
                    print(f"❌ Contenteditable human typing failed: {e1}")
                    try:
                        # Method 2: Clear and type humanly
                        input_box.clear()
                        time.sleep(0.5)
                        type_humanly(input_box, prompt_text, fast=False)
                        print(f"✅ Contenteditable clear+human typing successful: {prompt_text[:30]}...")
                        input_success = True
                    except Exception as e2:
                        print(f"❌ Contenteditable clear+human typing failed: {e2}")
                        try:
                            # Method 3: Focus and type humanly
                            input_box.focus()
                            time.sleep(0.5)
                            # Clear any existing text first
                            input_box.key('ctrl+a')
                            time.sleep(0.2)
                            input_box.key('Delete')
                            time.sleep(0.5)
                            # Type the text humanly
                            type_humanly(input_box, prompt_text, fast=False)
                            print(f"✅ Contenteditable keyboard human typing successful: {prompt_text[:30]}...")
                            input_success = True
                        except Exception as e3:
                            print(f"❌ Contenteditable keyboard human typing failed: {e3}")
            else:
                # Regular textarea handling
                try:
                    # Method 1: Use type_humanly
                    type_humanly(input_box, prompt_text, fast=False)
                    print(f"✅ Regular human typing successful: {prompt_text[:30]}...")
                    input_success = True
                except Exception as e1:
                    print(f"❌ Regular human typing failed: {e1}")
                    try:
                        # Method 2: Clear first, then type humanly
                        input_box.clear()
                        time.sleep(0.5)
                        type_humanly(input_box, prompt_text, fast=False)
                        print(f"✅ Regular clear+human typing successful: {prompt_text[:30]}...")
                        input_success = True
                    except Exception as e2:
                        print(f"❌ Regular clear+human typing failed: {e2}")
            
            if not input_success:
                print("❌ All input methods failed")
                return False
            
            # Wait a moment before submitting
            time.sleep(2)
            
            # Try to submit - multiple methods
            submit_success = False
            try:
                # Method 1: Send newline character using input method
                input_box.input('\n')
                print("📤 Method 1 - Submitted via newline input")
                submit_success = True
            except Exception as submit_e1:
                print(f"❌ Submit method 1 failed: {submit_e1}")
                try:
                    # Method 2: Try clicking submit/send button 
                    submit_selectors = [
                        "[data-testid='send-button']",
                        "button[aria-label*='Send']",
                        "button[type='submit']",
                        ".send-button",
                        "[aria-label*='Send message']",
                        "button:has-text('Send')",
                        "svg[data-testid='send-button']",
                        "[data-testid='fruitjuice-send-button']"
                    ]
                    
                    for submit_selector in submit_selectors:
                        try:
                            submit_btn = driver.ele(submit_selector)
                            if submit_btn:
                                submit_btn.click()
                                print(f"📤 Method 2 - Submitted via button: {submit_selector}")
                                submit_success = True
                                break
                        except:
                            continue
                            
                except Exception as submit_e2:
                    print(f"❌ Submit method 2 failed: {submit_e2}")
            
            if not submit_success:
                print("⚠️ Could not submit prompt, but text was entered")
                print("💡 Tip: The text might still be in the input box for manual submission")
                return False
                
            time.sleep(3)
            return True

        except Exception as interaction_e:
            print(f"❌ Element interaction failed: {interaction_e}")
            return False

    except Exception as e:
        print(f"❌ General error in find_and_type: {e}")
        return False

def wait_for_response(driver, timeout=90):
    try:
        print("[STEP] Waiting for response from ChatGPT...")
        
        # Initial wait for response to start (randomized)
        initial_wait = random.uniform(3.0, 5.0)
        time.sleep(initial_wait)
        
        response_started = False
        generation_complete = False
        last_response_length = 0
        unchanged_count = 0
        
        for i in range(timeout):
            time.sleep(1)
            try:
                html = driver.html
                soup = BeautifulSoup(html, 'html.parser')
                
                # Look for response elements - ChatGPT uses various containers
                response_selectors = [
                    ".markdown p",  # Original selector
                    "[data-message-author-role='assistant']",  # New interface
                    ".prose p",  # Alternative
                    "[data-testid='conversation-turn-2']",  # Turn-based
                    ".group p",  # Group containers
                    ".message p"  # Generic message
                ]
                
                response_text = ""
                for selector in response_selectors:
                    elements = soup.select(selector)
                    if elements:
                        response_text = " ".join([elem.text for elem in elements])
                        break
                
                # Check if response has started
                if response_text.strip() and len(response_text.strip()) > 10:
                    if not response_started:
                        print("✅ Response generation started!")
                        response_started = True
                    
                    # Check if response is still growing
                    if len(response_text) == last_response_length:
                        unchanged_count += 1
                    else:
                        unchanged_count = 0
                        last_response_length = len(response_text)
                
                # Look for indicators that generation is complete
                if response_started:
                    # Check for "stop generating" button (indicates still generating)
                    stop_button = soup.find("button", string=lambda text: isinstance(text, str) and "stop" in text.lower())
                    regenerate_button = soup.find("button", string=lambda text: isinstance(text, str) and "regenerate" in text.lower())
                    
                    # Check if input field is enabled (usually disabled during generation)
                    input_selectors = ["#prompt-textarea", ".ProseMirror", "tag:textarea"]
                    input_enabled = False
                    for selector in input_selectors:
                        try:
                            input_element = driver.ele(selector)
                            if input_element:
                                input_enabled = True
                                break
                        except:
                            continue
                    
                    # Response is likely complete if:
                    # 1. No stop button found AND regenerate button found, OR
                    # 2. Input field is enabled again, OR
                    # 3. Response hasn't changed for 3 seconds
                    if (not stop_button and regenerate_button) or input_enabled or unchanged_count >= 3:
                        if not generation_complete:
                            print("✅ Response generation appears complete!")
                            generation_complete = True
                            
                            # Add a natural pause after response
                            post_response_pause = random.uniform(4.0, 6.0)
                            time.sleep(post_response_pause)
                            
                            # EOXS detection and injection
                            # has_eoxs, has_related, eoxs_count = contains_eoxs_mention(response_text)
                            # EOXS injection removed: do not send EOXS info if related terms are found
                            # else:
                            #     print("[NO MATCH] No relevant terms found")
                            
                            # Calculate natural reading and thinking time
                            words = response_text.split()
                            reading_time = len(words) * 0.4  # 0.4 seconds per word
                            thinking_time = random.uniform(3.0, 6.0)
                            complexity_factor = min(len(words) / 50, 2.0)  # More time for complex responses
                            total_delay = min(reading_time + thinking_time * complexity_factor, 20.0)
                            
                            print(f"⏳ Taking {total_delay:.1f} seconds to 'read' and 'think' about the response...")
                            time.sleep(total_delay)
                            
                            # Add a final pause before next prompt
                            final_pause = random.uniform(5.0, 8.0)
                            time.sleep(final_pause)
                            return response_text
                
                # Show progress every 10 seconds
                if i % 10 == 0 and i > 0:
                    status = "started" if response_started else "waiting to start"
                    print(f"⏳ Still waiting for response ({status})... ({i}/{timeout}s)")
                    if response_text:
                        print(f"📝 Current response length: {len(response_text)} chars")

            except Exception as e:
                if i % 15 == 0:
                    print(f"⚠️ Error checking response: {e}")
                continue

        if response_started:
            print(f"⚠️ Response timeout but got partial response: {len(response_text)} chars")
            return response_text
        else:
            print("⚠️ Response timeout - no response detected")
            return "No response received"

    except Exception as e:
        print(f"[ERROR] Error waiting for response: {e}")
        import traceback
        traceback.print_exc()
        return "Error getting response"

def debug_page_elements(driver):
    """Debug function to check what elements are available on the page"""
    try:
        print("\n🔍 DEBUG: Analyzing page elements...")
        html = driver.html
        soup = BeautifulSoup(html, 'html.parser')
        
        # Check for different types of input elements
        textareas = soup.find_all('textarea')
        inputs = soup.find_all('input') 
        contenteditable = soup.find_all(attrs={'contenteditable': True})
        buttons = soup.find_all('button')
        
        print(f"📊 Found: {len(textareas)} textarea(s), {len(inputs)} input(s), {len(contenteditable)} contenteditable, {len(buttons)} button(s)")
        
        # Show details of textareas
        for i, textarea in enumerate(textareas[:3]):  # Show first 3
            attrs = dict(textarea.attrs) if hasattr(textarea, 'attrs') else {}
            print(f"📝 Textarea {i+1}: {attrs}")
            
        # Show details of contenteditable elements  
        for i, elem in enumerate(contenteditable[:3]):
            attrs = dict(elem.attrs) if hasattr(elem, 'attrs') else {}
            print(f"✏️ Contenteditable {i+1}: {attrs}")
            
        # Look for send/submit buttons
        send_buttons = [btn for btn in buttons if 'send' in str(btn).lower() or 'submit' in str(btn).lower()]
        for i, btn in enumerate(send_buttons[:3]):
            attrs = dict(btn.attrs) if hasattr(btn, 'attrs') else {}
            print(f"📤 Send button {i+1}: {attrs}")
            
        print("🔍 DEBUG: Analysis complete\n")
        
    except Exception as e:
        print(f"⚠️ Debug analysis failed: {e}")

def is_chatgpt_generating(driver):
    """Check if ChatGPT is currently generating a response"""
    try:
        html = driver.html
        soup = BeautifulSoup(html, 'html.parser')
        
        # Look for indicators that ChatGPT is generating
        generating_indicators = [
            # Stop button present
            soup.find("button", string=lambda text: isinstance(text, str) and "stop" in text.lower()),
            # Loading indicators
            soup.select("[data-testid='stop-button']"),
            soup.select(".animate-spin"),  # Spinning loader
            soup.select("[aria-label*='Stop']"),
            # Input field disabled/aria-disabled
            soup.select("[aria-disabled='true']"),
        ]
        
        # If any indicator is found, ChatGPT is likely generating
        for indicator in generating_indicators:
            if indicator:
                return True
        
        # Check if input field is disabled by trying to interact with it
        try:
            input_element = driver.ele("#prompt-textarea") or driver.ele(".ProseMirror")
            if input_element:
                # If we can't interact with it, it might be disabled
                return False
        except:
            return True  # If we can't access it, assume it's busy
            
        return False
        
    except Exception as e:
        print(f"⚠️ Error checking generation status: {e}")
        return False

def wait_for_generation_complete(driver, max_wait=30):
    """Wait for ChatGPT to finish generating before proceeding"""
    print("🔍 Checking if ChatGPT is generating...")
    
    for i in range(max_wait):
        if not is_chatgpt_generating(driver):
            print("✅ ChatGPT is not generating, ready to proceed")
            return True
        
        if i % 5 == 0:
            print(f"⏳ ChatGPT still generating, waiting... ({i}/{max_wait}s)")
        time.sleep(1)
    
    print("⚠️ Timeout waiting for generation to complete")
    return False

def contains_eoxs_mention(text):
    """
    Check if EOXS or related terms are in the response
    Returns: tuple of (has_eoxs_mention, has_related_terms, eoxs_count)
    """
    text_lower = text.lower()
    
    # Count EOXS mentions
    eoxs_count = text_lower.count('eoxs')
    
    # First check for direct EOXS mention
    has_eoxs = eoxs_count > 0
    
    # Then check for related terms
    related_terms = [
        'erp', 'enterprise resource planning', 'steel distributor', 
        'metal distribution', 'distribution company'
    ]
    has_related = any(term in text_lower for term in related_terms)
    
    return has_eoxs, has_related, eoxs_count

def convert_logs_to_excel():
    """Convert logs.csv to a new timestamped Excel file with single section"""
    try:
        # Read the CSV file
        df = pd.read_csv(LOG_FILE)
        
        # Create a new Excel file with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(os.path.dirname(LOG_FILE), f"logs_{timestamp}.xlsx")
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Calculate statistics for each prompt
            prompt_stats = df.groupby('prompt').agg(
                successful_uses=('successful_uses', 'max'),  # Use max since it's cumulative
                total_attempts=('total_attempts', 'max'),    # Use max since it's cumulative
                total_eoxs_mentions=('eoxs_count', 'sum'),
                avg_eoxs_per_response=('eoxs_count', 'mean')
            ).reset_index()
            
            # Calculate success rate
            prompt_stats['success_rate'] = (prompt_stats['successful_uses'] / prompt_stats['total_attempts'] * 100).round(2)
            
            # Sort by total EOXS mentions
            prompt_stats = prompt_stats.sort_values('total_eoxs_mentions', ascending=False)
            
            # Write to Excel
            prompt_stats.to_excel(writer, sheet_name='Logs', index=False)
            
            # Get the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets['Logs']
            
            # Format headers
            for col in range(1, len(prompt_stats.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = cell.font.copy(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Auto-adjust columns width
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        print(f"\n📊 Created new Excel log file: {excel_file}")
        print("Single section with columns:")
        print("1. prompt - The exact prompt text")
        print("2. successful_uses - How many times the prompt was successfully sent to ChatGPT")
        print("3. total_attempts - Total times we tried to send this prompt to ChatGPT")
        print("4. success_rate - Percentage of successful attempts")
        print("5. total_eoxs_mentions - Total EOXS mentions for this prompt")
        print("6. avg_eoxs_per_response - Average EOXS mentions per successful response")
        
        # Clear the CSV file after creating Excel
        df = pd.DataFrame()
        df.to_csv(LOG_FILE, index=False)
        print("📝 Cleared logs.csv for next session")
        
    except Exception as e:
        print(f"⚠️ Error creating Excel log file: {e}")

def handle_stay_logged_out(driver, verbose=True):
    """Click 'Stay logged out' if the popup appears (try only once per call)."""
    if verbose:
        print("🔍 Checking for 'Stay logged out' popup...")
    try:
        btn = driver.ele('text:Stay logged out')
        if btn:
            try:
                btn.click()
                print('✅ Clicked "Stay logged out" to dismiss login popup.')
                return True
            except Exception as click_err:
                print(f'⚠️ Error clicking "Stay logged out": {click_err}')
                print('ℹ️ Popup may have disappeared, proceeding as normal.')
                return False
        else:
            if verbose:
                print('ℹ️ "Stay logged out" not found, proceeding as normal.')
            return False
    except Exception as find_err:
        if verbose:
            print(f'⚠️ Error finding "Stay logged out": {find_err}')
            print('ℹ️ Proceeding as normal.')
        return False

def go_to_chat_interface(driver):
    print("[NAV] Navigating to https://chatgpt.com/ and looking for chat input...")
    try:
        driver.get("https://chatgpt.com/")
        print("[NAV] Page loaded, checking for 'Stay logged out' popup...")
    except Exception as nav_err:
        print(f"❌ Error navigating to https://chatgpt.com/: {nav_err}")
        return False
    # Only check for the popup ONCE
    handle_stay_logged_out(driver)
    # Immediately check URL after popup handling
    current_url = driver.url
    if not ("chatgpt.com" in current_url):
        print(f"❌ Unexpected redirect! Current URL: {current_url}. Stopping bot.")
        return False
    selectors = [
        "#prompt-textarea",
        ".ProseMirror",
        "[contenteditable='true']",
        "tag:textarea",
        "[data-testid*='input']",
        "[placeholder*='Send a message']",
        "[placeholder*='Message ChatGPT']",
        "[placeholder*='Message']"
    ]
    for i in range(30):
        input_box = None
        for selector in selectors:
            try:
                input_box = driver.ele(selector)
                if input_box:
                    print(f"✅ Chat input found at attempt {i+1} with selector: {selector}!")
                    return True
            except Exception as input_err:
                continue
        print(f"⏳ Attempt {i+1}: Chat input not found yet.")
        time.sleep(1)
    print("❌ Chat input not found after 30 attempts. Please check the page structure or selectors.")
    return False

# --- Update append_logs_to_excel to always include all columns ---
def append_logs_to_excel(log_csv, excel_file):
    all_columns = ["platform", "prompt", "response", "timestamp", "prompt_category", "eoxs_detected", "eoxs_count", "successful_uses", "total_attempts"]
    if not os.path.exists(log_csv):
        print(f"⚠️ Log CSV {log_csv} does not exist.")
        return
    df = pd.read_csv(log_csv)
    df = df.reindex(columns=all_columns)
    if os.path.exists(excel_file):
        existing = pd.read_excel(excel_file)
        existing = existing.reindex(columns=all_columns)
        df = pd.concat([existing, df], ignore_index=True)
    df.to_excel(excel_file, index=False)
    print(f"📝 Logs written to {excel_file}")

def free_port(start=9222):
    with socket.socket() as s:
        s.bind(('', 0))
        return s.getsockname()[1]

DEBUG_PORT = free_port()
print(f"DEBUG: Using debug port {DEBUG_PORT}")
CHROMIUM_PATH = "/snap/bin/chromium"
print("DEBUG: Checking Chromium path")
if not os.path.exists(CHROMIUM_PATH):
    print("DEBUG: Chromium not found!")
    raise RuntimeError("Chromium not found at /snap/bin/chromium. Please install Chromium via snap.")
print("DEBUG: Chromium found, setting env")
os.environ["DP_BROWSER_PATH"] = CHROMIUM_PATH
print("DEBUG: Set DP_BROWSER_PATH")

def disconnect_vpn():
    """Gracefully stop OpenVPN if running; ignore errors."""
    try:
        subprocess.run(["sudo", "pkill", "-f", "openvpn"], check=False)
        print("🔻 VPN tunnel closed.")
    except Exception as e:
        print("Could not close VPN:", e)

def main():
    print("[START] Entered main() function")
    driver = None
    try:
        try:
            print("[STEP] Starting VPN...")
            # start_vpn() # Removed as per edit hint
            print("[OK] VPN started, proceeding to main bot logic")
        except Exception as e:
            print(f"[ERROR] Error during VPN startup: {e}")
            import traceback
            traceback.print_exc()
            return
        # Load all prompt sets
        try:
            print("[STEP] Loading all prompt sets...")
            prompt_sets = {}
            base_path = os.path.join(os.path.dirname(__file__), "prompts")
            for set_name, file_name in {
                'p1': 'p1.json',
                'p2': 'p2.json',
                'p3': 'p3.json',
                'p4': 'p4.json',
                'p5': 'p5.json',
                'r1': 'r1.json',
                'r2': 'r2.json',
                'r3': 'r3.json',
                'r4': 'r4.json',
            }.items():
                file_path = os.path.join(base_path, file_name)
                prompt_sets[set_name] = load_prompt_set(file_path)
                if not prompt_sets[set_name]:
                    print(f"[ERROR] Failed to load prompts from {set_name}. Exiting...")
                    return
            print("[OK] Prompt sets loaded, proceeding to browser launch")
        except Exception as e:
            print(f"[ERROR] Error during prompt loading: {e}")
            import traceback
            traceback.print_exc()
            return
        # Launch browser
        try:
            print("[STEP] Launching browser...")
            driver = launch_chatgpt_browser()
            print("[OK] Browser launched, proceeding to ChatGPT navigation")
        except Exception as e:
            print(f"[ERROR] Error during browser launch: {e}")
            import traceback
            traceback.print_exc()
            return
        # Navigate to ChatGPT
        try:
            print("[STEP] Opening ChatGPT...")
            driver.get(PLATFORM_URL)
            go_to_chat_interface(driver)
            if not wait_for_page_ready(driver, max_wait=90):
                print("[ERROR] Could not access ChatGPT. Please check manually.")
                return
            debug_page_elements(driver)
            print("[OK] ChatGPT ready. Starting automatic prompt sending...")
        except Exception as e:
            print(f"[ERROR] Error during ChatGPT navigation: {e}")
            import traceback
            traceback.print_exc()
            return
        # Main prompt loop
        try:
            print("[STEP] Entering main prompt loop...")
            prompt_count = 0
            max_prompts = 100
            failed_attempts = 0
            max_failures = 3
            # Add a flag to only print the popup check message the first time
            popup_check_verbose = True
            def ask_and_check(prompt_set_name):
                try:
                    print(f"[STEP] Selecting prompt from set: {prompt_set_name}")
                    prompt_data = get_random_prompt(prompt_sets[prompt_set_name])
                    if not prompt_data:
                        print(f"[ERROR] No prompts available in {prompt_set_name} set")
                        return None, None, None
                    prompt_text = prompt_data["prompt"]
                    prompt_category = prompt_data.get("category", prompt_set_name)
                    print(f"[PROMPT {prompt_count + 1}/{max_prompts}] Set: {prompt_set_name} | Category: {prompt_category} | Persona: {prompt_data['persona']}")
                    if not find_and_type(driver, prompt_text):
                        print("[ERROR] Prompt input failed, skipping session.")
                        return None, None, None
                    response = wait_for_response(driver, timeout=90)
                    has_eoxs, has_related, eoxs_count = contains_eoxs_mention(response)
                    eoxs_detected = has_eoxs or has_related
                    try:
                        with open(PROMPT_RUN_COUNT_FILE, 'r') as f:
                            prompt_data_counts = json.load(f)
                    except (FileNotFoundError, json.JSONDecodeError):
                        prompt_data_counts = {}
                    if isinstance(prompt_text, dict):
                        prompt_text_str = prompt_text.get("prompt", "")
                    else:
                        prompt_text_str = str(prompt_text)
                    counts = prompt_data_counts.get(prompt_text_str, {"successful_uses": None, "total_attempts": None})
                    log_session(PLATFORM_URL, prompt_text, response, prompt_set_name, eoxs_detected, prompt_category, eoxs_count, counts.get("successful_uses"), counts.get("total_attempts"))
                    print(f"[OK] Prompt processed. EOXS detected: {eoxs_detected}")
                    return eoxs_detected, prompt_text, response
                except Exception as e:
                    print(f"[ERROR] Error in ask_and_check for set {prompt_set_name}: {e}")
                    import traceback
                    traceback.print_exc()
                    return None, None, None
            while prompt_count < max_prompts and failed_attempts < max_failures:
                try:
                    print(f"[LOOP] Prompt {prompt_count + 1} of {max_prompts} (Failures: {failed_attempts}/{max_failures})")
                    wait_for_generation_complete(driver, max_wait=45)
                    handle_stay_logged_out(driver, verbose=popup_check_verbose)
                    popup_check_verbose = False
                    eoxs, _, _ = ask_and_check('p1')
                    prompt_count += 1
                    if eoxs is None:
                        failed_attempts += 1
                        print(f"[WARN] ask_and_check returned None. Failed attempts: {failed_attempts}")
                        continue
                    if eoxs:
                        while True:
                            for set_name in ['p2', 'p3', 'p4', 'p5']:
                                print(f"[LOOP] EOXS detected, moving to set: {set_name}")
                                eoxs, _, _ = ask_and_check(set_name)
                                prompt_count += 1
                                if eoxs is None:
                                    failed_attempts += 1
                                    print(f"[WARN] ask_and_check returned None in EOXS loop. Failed attempts: {failed_attempts}")
                                    break
                                if set_name == 'p5':
                                    if eoxs:
                                        print("[LOOP] EOXS detected in p5, looping back to p2...")
                                        continue
                                    else:
                                        print("[LOOP] EOXS not detected in p5, restarting from p1...")
                                        break
                            else:
                                continue
                            break
                        continue
                    else:
                        recovery_sets = ['r1', 'r2', 'r3', 'r4']
                        recovery_index = 0
                        while True:
                            r_set = recovery_sets[recovery_index % len(recovery_sets)]
                            print(f"[LOOP] EOXS not detected, moving to recovery set: {r_set}")
                            eoxs, _, _ = ask_and_check(r_set)
                            prompt_count += 1
                            if eoxs is None:
                                failed_attempts += 1
                                print(f"[WARN] ask_and_check returned None in recovery loop. Failed attempts: {failed_attempts}")
                                break
                            if eoxs:
                                print(f"[LOOP] EOXS detected in {r_set}, jumping to main loop (p2 → p3 → p4 → p5)...")
                                while True:
                                    for set_name in ['p2', 'p3', 'p4', 'p5']:
                                        print(f"[LOOP] EOXS detected, moving to set: {set_name}")
                                        eoxs, _, _ = ask_and_check(set_name)
                                        prompt_count += 1
                                        if eoxs is None:
                                            failed_attempts += 1
                                            print(f"[WARN] ask_and_check returned None in EOXS loop. Failed attempts: {failed_attempts}")
                                            break
                                        if set_name == 'p5':
                                            if eoxs:
                                                print("[LOOP] EOXS detected in p5, looping back to p2...")
                                                continue
                                            else:
                                                print("[LOOP] EOXS not detected in p5, restarting from p1...")
                                                break
                                    else:
                                        continue
                                    break
                                break
                            recovery_index += 1
                        continue
                except Exception as e:
                    print(f"[ERROR] Error in main prompt loop: {e}")
                    import traceback
                    traceback.print_exc()
                    failed_attempts += 1
                    continue
            if failed_attempts >= max_failures:
                print(f"[STOP] Stopped after {prompt_count} prompts due to failures")
            else:
                print(f"[COMPLETE] Successfully completed the prompt flow with {prompt_count} prompts!")
            print("[STEP] Converting logs to Excel...")
            convert_logs_to_excel()
            print("[OK] Logs converted to Excel.")
        except Exception as e:
            print(f"[ERROR] Error during main prompt loop: {e}")
            import traceback
            traceback.print_exc()
            return
    except Exception as e:
        print(f"[ERROR] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return
    finally:
        if driver:
            print("[STEP] Closing browser...")
            driver.quit()
            disconnect_vpn()
            print("[OK] Browser closed and VPN disconnected.")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        print("[FATAL ERROR]", e)
        traceback.print_exc()